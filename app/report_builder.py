import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.audit_logger import get_connection

def get_data(query):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return pd.DataFrame(results)

def style_header(ws, row, col_count):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

def build_daily_dispatch_report():
    query = """
        SELECT
            d.dispatch_id,
            s.tracking_number,
            v.vendor_name,
            d.driver_name,
            d.vehicle_number,
            d.dispatch_time,
            d.delivery_time,
            d.tat_hours,
            d.delivery_status,
            d.sector
        FROM dispatch_logs d
        JOIN shipments s ON d.shipment_id = s.shipment_id
        JOIN vendors v ON s.vendor_id = v.vendor_id
        ORDER BY d.dispatch_time DESC
        LIMIT 500
    """
    df = get_data(query)
    if df.empty:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Dispatch MIS"

    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f"Daily Dispatch MIS Report — {datetime.now().strftime('%d %b %Y')}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    headers = ['Dispatch ID', 'Tracking No', 'Vendor', 'Driver',
               'Vehicle', 'Dispatch Time', 'Delivery Time', 'TAT Hours',
               'Status', 'Sector']
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
    style_header(ws, 2, len(headers))

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, row in enumerate(df.itertuples(), 3):
        values = list(row)[1:]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
            cell.alignment = Alignment(horizontal='center')
            if col_idx == 9:
                if str(value) == 'Delivered':
                    cell.fill = green_fill
                elif str(value) == 'Failed':
                    cell.fill = red_fill

    auto_width(ws)

    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "Summary"
    ws2['A1'].font = Font(bold=True, size=13, color="1F4E79")
    ws2['A2'] = "Total Dispatches"
    ws2['B2'] = len(df)
    ws2['A3'] = "Delivered"
    ws2['B3'] = len(df[df.iloc[:, 8] == 'Delivered']) if len(df.columns) > 8 else 0
    ws2['A4'] = "Failed"
    ws2['B4'] = len(df[df.iloc[:, 8] == 'Failed']) if len(df.columns) > 8 else 0
    ws2['A5'] = "Avg TAT Hours"
    ws2['B5'] = round(df.iloc[:, 7].astype(float).mean(), 2) if len(df.columns) > 7 else 0

    style_header(ws2, 1, 2)
    auto_width(ws2)

    os.makedirs('reports', exist_ok=True)
    filename = f"reports/Daily_Dispatch_MIS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def build_stock_summary_report():
    query = """
        SELECT
            sku_code,
            product_name,
            category,
            sector,
            SUM(quantity_in) as total_in,
            SUM(quantity_out) as total_out,
            SUM(current_stock) as current_stock,
            warehouse_location
        FROM inventory
        GROUP BY sku_code, product_name, category, sector, warehouse_location
        ORDER BY current_stock ASC
        LIMIT 500
    """
    df = get_data(query)
    if df.empty:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Summary"

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Weekly Stock Summary — {datetime.now().strftime('%d %b %Y')}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    headers = ['SKU Code', 'Product Name', 'Category', 'Sector',
               'Total In', 'Total Out', 'Current Stock', 'Location']
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
    style_header(ws, 2, len(headers))

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row_idx, row in enumerate(df.itertuples(), 3):
        values = list(row)[1:]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
            cell.alignment = Alignment(horizontal='center')
            if col_idx == 7:
                try:
                    stock = int(value)
                    if stock < 50:
                        cell.fill = red_fill
                    elif stock < 150:
                        cell.fill = yellow_fill
                except:
                    pass

    auto_width(ws)

    os.makedirs('reports', exist_ok=True)
    filename = f"reports/Stock_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def build_vendor_scorecard():
    query = """
        SELECT
            v.vendor_name,
            v.sector,
            v.city,
            v.rating,
            COUNT(s.shipment_id) as total_shipments,
            SUM(CASE WHEN s.status='Delivered' THEN 1 ELSE 0 END) as delivered,
            SUM(CASE WHEN s.status='Delayed' THEN 1 ELSE 0 END) as delayed,
            ROUND(AVG(d.tat_hours), 2) as avg_tat
        FROM vendors v
        LEFT JOIN shipments s ON v.vendor_id = s.vendor_id
        LEFT JOIN dispatch_logs d ON s.shipment_id = d.shipment_id
        GROUP BY v.vendor_id, v.vendor_name, v.sector, v.city, v.rating
        ORDER BY v.rating DESC
    """
    df = get_data(query)
    if df.empty:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Scorecard"

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Monthly Vendor Scorecard — {datetime.now().strftime('%B %Y')}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    headers = ['Vendor Name', 'Sector', 'City', 'Rating',
               'Total Shipments', 'Delivered', 'Delayed', 'Avg TAT (hrs)']
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
    style_header(ws, 2, len(headers))

    for row_idx, row in enumerate(df.itertuples(), 3):
        values = list(row)[1:]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
            cell.alignment = Alignment(horizontal='center')

    auto_width(ws)

    os.makedirs('reports', exist_ok=True)
    filename = f"reports/Vendor_Scorecard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename